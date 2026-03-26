"""yuclaw/output/excel.py — Professional financial spreadsheet output."""
from __future__ import annotations
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY  = "0A1B36"
GOLD  = "C8981F"
LIGHT = "EFF3FA"
GREEN = "0A6040"
RED   = "A81818"


def _thin(): return Side(style="thin", color="B8C8DC")
def _border(): return Border(left=_thin(), right=_thin(), top=_thin(), bottom=_thin())


class ExcelExporter:

    def export_research(self, ticker: str, analysis: dict, output_path: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{ticker} Research"

        # Title row
        ws.merge_cells("A1:G1")
        ws["A1"] = f"YUCLAW — {ticker} Investment Analysis — {datetime.now().strftime('%Y-%m-%d')}"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
        ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        row = 3

        # Summary
        ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, color="FFFFFF", name="Arial")
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
        ws.merge_cells(f"A{row}:G{row}")
        row += 1
        summary_cell = ws.cell(row=row, column=1, value=analysis.get("summary", ""))
        summary_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:G{row}")
        ws.row_dimensions[row].height = 60
        row += 2

        # Thesis
        if analysis.get("thesis"):
            ws.cell(row=row, column=1, value="THESIS").font = Font(bold=True, color="FFFFFF", name="Arial")
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
            ws.merge_cells(f"A{row}:G{row}")
            row += 1
            for scenario, text in analysis["thesis"].items():
                ws.cell(row=row, column=1, value=scenario.upper()).font = Font(bold=True, name="Arial")
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT)
                c = ws.cell(row=row, column=2, value=text)
                c.alignment = Alignment(wrap_text=True)
                ws.merge_cells(f"B{row}:G{row}")
                ws.row_dimensions[row].height = 40
                row += 1
            row += 1

        # Key Metrics
        if analysis.get("key_metrics"):
            ws.cell(row=row, column=1, value="KEY METRICS").font = Font(bold=True, color="FFFFFF", name="Arial")
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=GOLD)
            ws.merge_cells(f"A{row}:G{row}")
            row += 1
            for i, m in enumerate(analysis["key_metrics"]):
                bg = LIGHT if i % 2 == 0 else "FFFFFF"
                ws.cell(row=row, column=1, value=m.get("name","")).font = Font(bold=True, name="Arial")
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=bg)
                ws.cell(row=row, column=2, value=m.get("value","")).fill = PatternFill("solid", fgColor=bg)
                ws.cell(row=row, column=3, value=f"pg.{m.get('page','')}").fill = PatternFill("solid", fgColor=bg)
                ws.cell(row=row, column=3).font = Font(color="888888", italic=True, name="Arial")
                sig = ws.cell(row=row, column=4, value=m.get("significance",""))
                sig.fill = PatternFill("solid", fgColor=bg)
                ws.merge_cells(f"D{row}:G{row}")
                row += 1
            row += 1

        # Assumptions
        if analysis.get("key_assumptions"):
            ws.cell(row=row, column=1, value="KEY ASSUMPTIONS").font = Font(bold=True, color="FFFFFF", name="Arial")
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
            ws.merge_cells(f"A{row}:G{row}")
            row += 1
            for a in analysis["key_assumptions"]:
                ws.cell(row=row, column=1, value="▸")
                c = ws.cell(row=row, column=2, value=a)
                c.alignment = Alignment(wrap_text=True)
                ws.merge_cells(f"B{row}:G{row}")
                ws.row_dimensions[row].height = 30
                row += 1
            row += 1

        # Risks
        if analysis.get("risks"):
            ws.cell(row=row, column=1, value="RISKS").font = Font(bold=True, color="FFFFFF", name="Arial")
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=RED)
            ws.merge_cells(f"A{row}:G{row}")
            row += 1
            for r_text in analysis["risks"]:
                ws.cell(row=row, column=1, value="⚠")
                c = ws.cell(row=row, column=2, value=r_text)
                c.alignment = Alignment(wrap_text=True)
                ws.merge_cells(f"B{row}:G{row}")
                ws.row_dimensions[row].height = 28
                row += 1

        # Evidence count footer
        row += 1
        n_evidence = len(analysis.get("evidence_node_ids", []))
        ws.cell(row=row, column=1, value=f"Evidence anchors: {n_evidence} claims traceable to source documents")
        ws.cell(row=row, column=1).font = Font(italic=True, color="888888", name="Arial")
        ws.merge_cells(f"A{row}:G{row}")

        # Column widths
        widths = [4, 28, 12, 8, 20, 20, 20]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        wb.save(output_path)
        return output_path

    def export_validation_report(self, result: dict, output_path: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation Report"

        ws.merge_cells("A1:D1")
        ws["A1"] = f"YUCLAW Adversarial Validation — {result.get('strategy_id','')}"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
        fill_color = GREEN if result.get("passed") else RED
        ws["A1"].fill = PatternFill("solid", fgColor=fill_color)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        row = 3
        fields = [
            ("Verdict",           result.get("verdict","")),
            ("Calmar Ratio",      result.get("calmar_ratio","")),
            ("Survival Rate",     result.get("survival_rate","")),
            ("Scenarios Tested",  result.get("scenarios_tested","")),
            ("Scenarios Survived",result.get("scenarios_survived","")),
            ("Scenarios Killed",  result.get("scenarios_killed","")),
            ("Worst Drawdown",    result.get("worst_drawdown","")),
            ("Status",            result.get("status","")),
        ]
        for label, value in fields:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, name="Arial")
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT)
            ws.cell(row=row, column=2, value=str(value))
            ws.merge_cells(f"B{row}:D{row}")
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="FATAL SCENARIOS").font = Font(bold=True, color="FFFFFF", name="Arial")
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=RED)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1
        for s in result.get("fatal_scenarios", []):
            ws.cell(row=row, column=1, value="✗")
            c = ws.cell(row=row, column=2, value=s)
            c.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"B{row}:D{row}")
            ws.row_dimensions[row].height = 40
            row += 1

        for col, w in [(1,5),(2,60),(3,20),(4,20)]:
            ws.column_dimensions[get_column_letter(col)].width = w

        wb.save(output_path)
        return output_path
